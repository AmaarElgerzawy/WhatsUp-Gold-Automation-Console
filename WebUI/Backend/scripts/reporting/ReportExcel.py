import calendar
import json
import os
import pyodbc
from datetime import datetime, timedelta, timezone
from openpyxl.utils import get_column_letter
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import resend
from constants import (
    get_connection_string,
    REPORT_SCHEDULE_JSON_FILE,
)
try:
    # When imported as part of the FastAPI app (scripts.reporting package)
    from scripts.reporting.DeviceUpTimeReport import run_sp_group_device_uptime, write_excel as write_uptime_excel
except ImportError:
    # Fallback for running this file directly
    from DeviceUpTimeReport import run_sp_group_device_uptime, write_excel as write_uptime_excel


# This should match your report folder from the other script
OUTPUT_FOLDER = r"C:\WUG_Exports"
# Preferred JSON-based schedule shared with the FastAPI API
SCHEDULE_JSON_FILE = str(REPORT_SCHEDULE_JSON_FILE)

# Resend configuration (provided by user)
RESEND_API_KEY = "re_j4PGF4Ev_Jr1M8fDaaARvJS7WocdFACDt"
RESEND_FROM = "onboarding@resend.dev"
RESEND_TO = "snipergolden1234@gmail.com"

resend.api_key = RESEND_API_KEY
STATE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "state.json")
# =================================
def send_all_reports_via_email(start_date, end_date):
    """
    Attach all XLSX reports from OUTPUT_FOLDER and send in one email
    using Resend.
    """
    folder = Path(OUTPUT_FOLDER)
    if not folder.exists():
        print(f"Report folder does not exist: {folder}")
        return

    # Pick all Excel reports we generated, e.g. ActiveMonitorAvailability_*.xlsx
    report_files = sorted(folder.glob("ActiveMonitorAvailability_*.xlsx"))

    if not report_files:
        print("No report files found to send.")
        return

    subject = f"WUG Monthly Reports: {start_date:%Y-%m-%d} to {end_date:%Y-%m-%d}"
    html_body = (
        f"<p>Attached are the Active Monitor Availability reports for all groups<br>"
        f"from {start_date:%Y-%m-%d %H:%M} to {end_date:%Y-%m-%d %H:%M}.</p>"
        f"<p>Total reports: {len(report_files)}</p>"
    )

    attachments = []
    for report_path in report_files:
        try:
            with open(report_path, "rb") as f:
                content_bytes = f.read()
            attachment: resend.Attachment = {
                "content": list(content_bytes),
                "filename": os.path.basename(report_path),
            }
            attachments.append(attachment)
        except Exception as e:
            print(f"[EMAIL] Failed to read attachment {report_path}: {e}")

    params: resend.Emails.SendParams = {
        "from": RESEND_FROM,
        "to": [RESEND_TO],
        "subject": subject,
        "html": html_body,
    }
    if attachments:
        params["attachments"] = attachments

    try:
        resend.Emails.send(params)
        print(f"[EMAIL] Resend monthly email sent with {len(attachments)} attachment(s).")
    except Exception as e:
        print(f"[EMAIL] ERROR sending monthly email via Resend: {e}")

# ========== CONFIG ==========
WEB_USER_ID = 1  # same web user as in WUG (nWebUserID)
OUTPUT_FOLDER = r"C:\WUG_Exports"
SQL_PERCENT_MULTIPLIER = 1000000000.0
REPORT_PERCENT_DIVISOR = 10000000.0
ROUND_TO_PLACES_EXPORT = 7
# ============================
def send_single_report_email(group_name, report_path, start_date, end_date):
    """
    Send a single report as an attachment using Resend.
    """
    subject = f"WUG Report for {group_name}: {start_date:%Y-%m-%d %H:%M} → {end_date:%Y-%m-%d %H:%M}"
    html_body = (
        f"<p>Attached is the report for group '<strong>{group_name}</strong>'.</p>"
        f"<p>Period: {start_date:%Y-%m-%d %H:%M} to {end_date:%Y-%m-%d %H:%M}.</p>"
    )

    attachments = []
    try:
        with open(report_path, "rb") as f:
            content_bytes = f.read()
        attachment: resend.Attachment = {
            "content": list(content_bytes),
            "filename": os.path.basename(report_path),
        }
        attachments.append(attachment)
    except Exception as e:
        print(f"[EMAIL] Failed to read attachment {report_path}: {e}")

    params: resend.Emails.SendParams = {
        "from": RESEND_FROM,
        "to": [RESEND_TO],
        "subject": subject,
        "html": html_body,
    }
    if attachments:
        params["attachments"] = attachments

    try:
        resend.Emails.send(params)
        print(f"[EMAIL] Resend email sent for group {group_name} with attachment {report_path}")
    except Exception as e:
        print(f"[EMAIL] ERROR sending email via Resend for group {group_name}: {e}")

def load_state():
    if not os.path.exists(STATE_FILE):
        return {}
    with open(STATE_FILE, "r") as f:
        return json.load(f)

def save_state(state):
    with open(STATE_FILE, "w") as f:
        json.dump(state, f)

# ---------- Simple scheduler: weekly or monthly ----------
# Use consistent time: set TZ env for server so "run_time" is unambiguous (or use UTC below).
def _scheduler_now() -> datetime:
    return datetime.now()

def _parse_time(s: str):
    """Parse 'HH:MM' or 'HH:MM:SS' -> (hour, minute)."""
    s = (s or "00:00").strip()
    parts = s.split(":")
    h = int(parts[0]) if parts else 0
    m = int(parts[1]) if len(parts) > 1 else 0
    return h, m

def _weekday_index(day_code: str) -> int:
    """Map 'mon'..'sun' to Python weekday (Mon=0)."""
    code = (day_code or "").strip().lower()
    mapping = {"mon": 0, "tue": 1, "wed": 2, "thu": 3, "fri": 4, "sat": 5, "sun": 6}
    return mapping.get(code, 0)

def _trigger_fires(job: dict, now: datetime, last_run_iso: str) -> tuple[bool, datetime | None]:
    """
    Simple trigger: weekly = today is run_day and time >= run_time; monthly = today is run_day_of_month and time >= run_time.
    Returns (should_run, anchor_datetime).
    """
    schedule_type = (job.get("type") or "weekly").strip().lower()
    run_time_str = job.get("run_time") or "00:00"
    h, m = _parse_time(run_time_str)
    anchor = now.replace(hour=h, minute=m, second=0, microsecond=0)

    if schedule_type == "monthly":
        try:
            day_of_month = int(job.get("run_day_of_month") or 0)
        except (ValueError, TypeError):
            return False, None
        if day_of_month < 1 or now.day != day_of_month:
            return False, None
        if now < anchor:
            return False, anchor
        if last_run_iso:
            try:
                last_run = datetime.fromisoformat(last_run_iso)
                if last_run >= anchor:
                    return False, anchor
            except Exception:
                pass
        return True, anchor

    # weekly: run_day = weekday (mon..sun)
    run_day = job.get("run_day")
    if not run_day:
        return False, None
    wd = _weekday_index(str(run_day))
    if now.weekday() != wd:
        return False, None
    if now < anchor:
        return False, anchor
    if last_run_iso:
        try:
            last_run = datetime.fromisoformat(last_run_iso)
            if last_run >= anchor:
                return False, anchor
        except Exception:
            pass
    return True, anchor

def _compute_window_from_trigger(anchor_dt: datetime, job: dict) -> tuple[datetime, datetime]:
    """
    Compute report data window.
    - Weekly: last week (week containing run date). period_start_day/weekday at period_start_time -> period_end_day/weekday at period_end_time.
    - Monthly: previous month. period_start_day (1-31) at period_start_time -> period_end_day at period_end_time.
    """
    schedule_type = (job.get("type") or "weekly").strip().lower()
    run_date = anchor_dt.date()

    if schedule_type == "monthly":
        if run_date.month == 1:
            prev_first = run_date.replace(year=run_date.year - 1, month=12, day=1)
        else:
            prev_first = run_date.replace(month=run_date.month - 1, day=1)
        _, last_day = calendar.monthrange(prev_first.year, prev_first.month)
        start_d = min(int(job.get("period_start_day") or 1), last_day)
        end_d = min(int(job.get("period_end_day") or last_day), last_day)
        start_d = max(1, start_d)
        end_d = max(start_d, end_d)
        sh, sm = _parse_time(job.get("period_start_time") or "00:00")
        eh, em = _parse_time(job.get("period_end_time") or "23:59")
        start_dt = datetime(prev_first.year, prev_first.month, start_d, sh, sm, 0, 0)
        end_dt = datetime(prev_first.year, prev_first.month, end_d, eh, em, 59, 999999)
        return start_dt, end_dt

    # weekly: same week as run_date (Mon–Sun containing the run day). period_*_day = weekday code (mon..sun).
    week_start = run_date - timedelta(days=run_date.weekday())
    start_wd = _weekday_index(str(job.get("period_start_day") or "mon"))
    end_wd = _weekday_index(str(job.get("period_end_day") or "wed"))
    start_date = week_start + timedelta(days=start_wd)
    end_date = week_start + timedelta(days=end_wd)
    sh, sm = _parse_time(job.get("period_start_time") or "00:00")
    eh, em = _parse_time(job.get("period_end_time") or "23:59")
    start_dt = datetime(start_date.year, start_date.month, start_date.day, sh, sm, 0, 0)
    end_dt = datetime(end_date.year, end_date.month, end_date.day, eh, em, 59, 999999)
    return start_dt, end_dt


def load_schedule_config():
    """
    Load schedule from report_schedule.json. Simple two-mode model:
    - type: "weekly" | "monthly", group, report_type: "availability" | "uptime" | "both"
    - Weekly: run_day, run_time, period_start_day, period_start_time, period_end_day, period_end_time
    - Monthly: run_day_of_month, run_time, period_start_day, period_start_time, period_end_day, period_end_time
    """
    if not os.path.exists(SCHEDULE_JSON_FILE):
        return []

    with open(SCHEDULE_JSON_FILE, "r", encoding="utf-8") as f:
        raw = json.load(f)

    rows = raw.get("rows", [])
    if not isinstance(rows, list) or not rows:
        return []

    def _clean(v):
        if v is None:
            return None
        s = str(v).strip()
        return s if s else None

    jobs = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        job_id = _clean(row.get("id"))
        group_name = _clean(row.get("group"))
        if not group_name:
            continue

        schedule_type = _clean(row.get("type")) or "weekly"
        report_type = _clean(row.get("report_type")) or "both"
        entry = {
            "id": job_id,
            "group": group_name,
            "type": schedule_type,
            "report_type": report_type,
            "run_day": _clean(row.get("run_day")),
            "run_time": _clean(row.get("run_time")),
            "period_start_day": row.get("period_start_day"),
            "period_start_time": _clean(row.get("period_start_time")),
            "period_end_day": row.get("period_end_day"),
            "period_end_time": _clean(row.get("period_end_time")),
            "run_day_of_month": row.get("run_day_of_month"),
        }
        if schedule_type == "monthly":
            try:
                entry["run_day_of_month"] = int(entry.get("run_day_of_month") or 1)
            except (ValueError, TypeError):
                entry["run_day_of_month"] = 1
            for k in ("period_start_day", "period_end_day"):
                try:
                    entry[k] = int(entry.get(k) or 1)
                except (ValueError, TypeError):
                    entry[k] = 1

        entry = {k: v for k, v in entry.items() if v is not None}
        if "id" not in entry:
            entry["id"] = None
        jobs.append(entry)

    return jobs

def get_duration_from_seconds(total_seconds: int) -> str:
    """Approximate of GetDurationFromSeconds from ASP code."""
    if total_seconds is None or total_seconds < 0:
        total_seconds = 0

    total_seconds = int(total_seconds)
    minutes, _ = divmod(total_seconds, 60)
    hours, minutes = divmod(minutes, 60)
    days, hours = divmod(hours, 24)

    parts = []
    if days:
        parts.append(f"{days}d")
    if hours:
        parts.append(f"{hours}h")
    if minutes or not parts:
        parts.append(f"{minutes}m")

    return " ".join(parts)

def get_active_monitor_availability(device_group_id: int,
                                    start_date: datetime,
                                    end_date: datetime):
    """
    Runs the same SQL you just tested in SSMS (dynamic LoadDeviceGroup + main aggregation),
    and returns list of rows ready to export.
    """

    # Format dates like '2025-11-23T00:00:00'
    start_str = start_date.strftime("%Y-%m-%dT%H:%M:%S")
    end_str = end_date.strftime("%Y-%m-%dT%H:%M:%S")

    # This is basically your working SSMS script,
    # but with {group_id}, {web_user_id}, {start}, {end} filled in by Python.
    sql_template = f"""
DECLARE @DeviceGroupID INT;
DECLARE @WebUserID INT;
DECLARE @StartDate DATETIME;
DECLARE @EndDate DATETIME;

SET @DeviceGroupID = {device_group_id};
SET @WebUserID = {WEB_USER_ID};
SET @StartDate = '{start_str}';
SET @EndDate = '{end_str}';

DECLARE @PivotTableName SYSNAME;
SET @PivotTableName = 'PivotDeviceToGroupTemp';

IF OBJECT_ID(@PivotTableName) IS NOT NULL
BEGIN
    DECLARE @dropSql NVARCHAR(4000);
    SET @dropSql = N'DROP TABLE ' + QUOTENAME(@PivotTableName) + N';';
    EXEC (@dropSql);
END;

EXEC LoadDeviceGroup @DeviceGroupID, @WebUserID, @PivotTableName;

DECLARE @sql NVARCHAR(MAX);

SET @sql = N'
SELECT
    Device.nDeviceID,
    Device.sDisplayName,
    NetworkInterface.nNetworkInterfaceID,
    NetworkInterface.sNetworkName,
    NetworkInterface.sNetworkAddress,
    Device.nWorstStateID,
    Device.nBestStateID,
    ActiveMonitorType.sMonitorTypeName,
    PivotActiveMonitorTypeToDevice.sArgument,
    PivotActiveMonitorTypeToDevice.nPivotActiveMonitorTypeToDeviceID,
    PivotActiveMonitorTypeToDevice.sComment,
    CAST(Device.sNote AS NVARCHAR(MAX)) AS sNote,

    -- seconds in each state
    SUM(
        CASE MonitorState.nInternalMonitorState WHEN 1 THEN
            DATEDIFF(
                SECOND,
                CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
                ISNULL(
                    CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                    CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
                )
            )
        ELSE 0 END
    ) AS nDownSeconds,

    SUM(
        CASE MonitorState.nInternalMonitorState WHEN 2 THEN
            DATEDIFF(
                SECOND,
                CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
                ISNULL(
                    CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                    CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
                )
            )
        ELSE 0 END
    ) AS nMaintenanceSeconds,

    SUM(
        CASE MonitorState.nInternalMonitorState WHEN 3 THEN
            DATEDIFF(
                SECOND,
                CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
                ISNULL(
                    CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                    CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
                )
            )
        ELSE 0 END
    ) AS nUpSeconds,

    SUM(
        CASE MonitorState.nInternalMonitorState WHEN -1 THEN
            DATEDIFF(
                SECOND,
                CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
                ISNULL(
                    CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                    CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
                )
            )
        ELSE 0 END
    ) AS nUnknownSeconds,

    -- total seconds (any state)
    SUM(
        DATEDIFF(
            SECOND,
            CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
            ISNULL(
                CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
            )
        )
    ) AS nTotalSeconds,

    -- Percentages scaled like WUG: * 1000000000
    (1.0 * SUM(
        CASE MonitorState.nInternalMonitorState WHEN 1 THEN
            DATEDIFF(
                SECOND,
                CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
                ISNULL(
                    CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                    CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
                )
            )
        ELSE 0 END
    ) / NULLIF(
        SUM(
            DATEDIFF(
                SECOND,
                CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
                ISNULL(
                    CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                    CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
                )
            )
        ), 0)
    ) * 1000000000.0 AS nDownPercent,

    (1.0 * SUM(
        CASE MonitorState.nInternalMonitorState WHEN 2 THEN
            DATEDIFF(
                SECOND,
                CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
                ISNULL(
                    CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                    CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
                )
            )
        ELSE 0 END
    ) / NULLIF(
        SUM(
            DATEDIFF(
                SECOND,
                CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
                ISNULL(
                    CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                    CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
                )
            )
        ), 0)
    ) * 1000000000.0 AS nMaintenancePercent,

    (1.0 * SUM(
        CASE MonitorState.nInternalMonitorState WHEN 3 THEN
            DATEDIFF(
                SECOND,
                CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
                ISNULL(
                    CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                    CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
                )
            )
        ELSE 0 END
    ) / NULLIF(
        SUM(
            DATEDIFF(
                SECOND,
                CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
                ISNULL(
                    CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                    CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
                )
            )
        ), 0)
    ) * 1000000000.0 AS nUpPercent,

    (1.0 * SUM(
        CASE MonitorState.nInternalMonitorState WHEN -1 THEN
            DATEDIFF(
                SECOND,
                CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
                ISNULL(
                    CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                    CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
                )
            )
        ELSE 0 END
    ) / NULLIF(
        SUM(
            DATEDIFF(
                SECOND,
                CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
                ISNULL(
                    CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                    CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
                )
            )
        ), 0)
    ) * 1000000000.0 AS nUnknownPercent

FROM ActiveMonitorStateChangeLog
INNER JOIN MonitorState
    ON MonitorState.nMonitorStateID = ActiveMonitorStateChangeLog.nMonitorStateID
INNER JOIN PivotActiveMonitorTypeToDevice
    ON PivotActiveMonitorTypeToDevice.nPivotActiveMonitorTypeToDeviceID = ActiveMonitorStateChangeLog.nPivotActiveMonitorTypeToDeviceID
INNER JOIN Device
    ON Device.nDeviceID = PivotActiveMonitorTypeToDevice.nDeviceID
INNER JOIN NetworkInterface
    ON NetworkInterface.nNetworkInterfaceID = Device.nDefaultNetworkInterfaceID
INNER JOIN ActiveMonitorType
    ON ActiveMonitorType.nActiveMonitorTypeID = PivotActiveMonitorTypeToDevice.nActiveMonitorTypeID
INNER JOIN ' + QUOTENAME(@PivotTableName) + N'
    ON Device.nDeviceID = ' + QUOTENAME(@PivotTableName) + N'.nDeviceID
WHERE
    ISNULL(ActiveMonitorType.bRemoved,0) <> 1
    AND ISNULL(Device.bRemoved,0)      <> 1
    AND (
        (dStartTime >= @StartDate AND ISNULL(dEndTime, GETDATE()) <= @EndDate)
        OR (dStartTime <= @EndDate   AND ISNULL(dEndTime, GETDATE()) >= @StartDate)
        OR (dStartTime <= @EndDate   AND ISNULL(dEndTime, GETDATE()) >= @EndDate)
        OR (dStartTime <= @StartDate AND ISNULL(dEndTime, GETDATE()) >= @EndDate)
    )
GROUP BY
    Device.nDeviceID,
    Device.sDisplayName,
    NetworkInterface.nNetworkInterfaceID,
    NetworkInterface.sNetworkName,
    NetworkInterface.sNetworkAddress,
    Device.nWorstStateID,
    Device.nBestStateID,
    PivotActiveMonitorTypeToDevice.nPivotActiveMonitorTypeToDeviceID,
    PivotActiveMonitorTypeToDevice.sArgument,
    PivotActiveMonitorTypeToDevice.sComment,
    ActiveMonitorType.sMonitorTypeName,
    CAST(Device.sNote AS NVARCHAR(MAX))
ORDER BY
    Device.sDisplayName ASC,
    ActiveMonitorType.sMonitorTypeName ASC;
';

EXEC sp_executesql
    @sql,
    N'@StartDate datetime, @EndDate datetime',
    @StartDate = @StartDate,
    @EndDate   = @EndDate;

IF OBJECT_ID(@PivotTableName) IS NOT NULL
BEGIN
    DECLARE @dropSql2 NVARCHAR(4000);
    SET @dropSql2 = N'DROP TABLE ' + QUOTENAME(@PivotTableName) + N';';
    EXEC (@dropSql2);
END;
"""

    conn = pyodbc.connect(get_connection_string())
    cursor = conn.cursor()

    cursor.execute(sql_template)

    rows = cursor.fetchall()
    cols = [c[0] for c in cursor.description]

    results = []

    for row in rows:
        rec = dict(zip(cols, row))

        def scaled(p):
            if p is None:
                return 0.0
            return float(p) / REPORT_PERCENT_DIVISOR

        up_pct = scaled(rec["nUpPercent"])
        down_pct = scaled(rec["nDownPercent"])
        maint_pct = scaled(rec["nMaintenancePercent"])
        unknown_pct = scaled(rec["nUnknownPercent"])

        up_dur = get_duration_from_seconds(rec["nUpSeconds"])
        down_dur = get_duration_from_seconds(rec["nDownSeconds"])
        maint_dur = get_duration_from_seconds(rec["nMaintenanceSeconds"])
        unknown_dur = get_duration_from_seconds(rec["nUnknownSeconds"])
        total_dur = get_duration_from_seconds(rec["nTotalSeconds"])

        sMonitorTypeName = rec["sMonitorTypeName"]
        sArgument = rec["sArgument"] or ""
        sComment = rec["sComment"] or ""
        
        sIPAddress = rec["sNetworkAddress"]
        sNote = rec["sNote"]

        monitor_name = sMonitorTypeName
        if sArgument:
            monitor_name += f" ({sArgument})"
        if sComment:
            monitor_name += f" - {sComment}"

        results.append({
            "Device": rec["sDisplayName"],
            "IPAddress": sIPAddress,
            "Note": sNote,
            "Monitor": monitor_name,
            "UpPercent": round(up_pct, ROUND_TO_PLACES_EXPORT),
            "UpDuration": up_dur,
            "MaintenancePercent": round(maint_pct, ROUND_TO_PLACES_EXPORT),
            "MaintenanceDuration": maint_dur,
            "UnknownPercent": round(unknown_pct, ROUND_TO_PLACES_EXPORT),
            "UnknownDuration": unknown_dur,
            "DownPercent": round(down_pct, ROUND_TO_PLACES_EXPORT),
            "DownDuration": down_dur,
            "TotalDuration": total_dur,
        })

    cursor.close()
    conn.close()
    return results

def write_excel_for_group(device_group_id: int,
                          start_date: datetime,
                          end_date: datetime,
                          group_name: str = None):
    rows = get_active_monitor_availability(device_group_id, start_date, end_date)
    if not rows:
        print(f"No data for group {device_group_id}")
        return

    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    if not group_name:
        group_name = f"group_{device_group_id}"

    safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in group_name)
    filename = f"ActiveMonitorAvailability_{safe_name}.xlsx"
    path = os.path.join(OUTPUT_FOLDER, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Active Monitor Availability"

    # Columns exactly like the WUG export
    headers = [
        "Device",
        "Monitor",
        "Up",
        "Up Duration",
        "Maintenance",
        "Maintenance Duration",
        "Unknown",
        "Unknown Duration",
        "Down",
        "Down Duration",
        "Total Duration",
        "Network Address (IP)",
        "Note",
    ]
    last_col_letter = get_column_letter(len(headers))

    # Common styles
    title_font = Font(size=14, bold=True)
    subtitle_font = Font(size=11)
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center")
    right_align = Alignment(horizontal="right")
    header_fill = PatternFill("solid", fgColor="C0C0C0")  # light grey like your screenshot
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ----- Row 1: Title -----
    ws["A1"].value = "Active Monitor Availability"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align
    ws.merge_cells(f"A1:{last_col_letter}1")

    # ----- Row 2: Subtitle (Group - start - end) -----
    start_long = start_date.strftime("%A, %B %d, %Y %I:%M:%S %p")
    end_long = end_date.strftime("%A, %B %d, %Y %I:%M:%S %p")
    ws["A2"].value = f"{group_name} - {start_long} - {end_long}"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = center_align
    ws.merge_cells(f"A2:{last_col_letter}2")

    # ----- Row 3: Headers -----
    header_row_idx = 3
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # ----- Data rows -----
    row_idx = header_row_idx + 1

    for r in rows:
        up_fraction = (r["UpPercent"] or 0.0) / 100.0
        maint_fraction = (r["MaintenancePercent"] or 0.0) / 100.0
        unknown_fraction = (r["UnknownPercent"] or 0.0) / 100.0
        down_fraction = (r["DownPercent"] or 0.0) / 100.0

        # Device
        c = ws.cell(row=row_idx, column=1, value=r["Device"])
        c.border = thin_border
        
        # Monitor
        c = ws.cell(row=row_idx, column=2, value=r["Monitor"])
        c.border = thin_border

        # Up %
        up_cell = ws.cell(row=row_idx, column=3, value=up_fraction)
        up_cell.number_format = "0.0000000%"
        up_cell.alignment = right_align
        up_cell.border = thin_border

        # Up Duration
        c = ws.cell(row=row_idx, column=4, value=r["UpDuration"])
        c.alignment = right_align
        c.border = thin_border

        # Maintenance %
        maint_cell = ws.cell(row=row_idx, column=5, value=maint_fraction)
        maint_cell.number_format = "0.0000000%"
        maint_cell.alignment = right_align
        maint_cell.border = thin_border

        # Maintenance Duration
        c = ws.cell(row=row_idx, column=6, value=r["MaintenanceDuration"])
        c.alignment = right_align
        c.border = thin_border

        # Unknown %
        unknown_cell = ws.cell(row=row_idx, column=7, value=unknown_fraction)
        unknown_cell.number_format = "0.0000000%"
        unknown_cell.alignment = right_align
        unknown_cell.border = thin_border

        # Unknown Duration
        c = ws.cell(row=row_idx, column=8, value=r["UnknownDuration"])
        c.alignment = right_align
        c.border = thin_border

        # Down %
        down_cell = ws.cell(row=row_idx, column=9, value=down_fraction)
        down_cell.number_format = "0.0000000%"
        down_cell.alignment = right_align
        down_cell.border = thin_border

        # Down Duration
        c = ws.cell(row=row_idx, column=10, value=r["DownDuration"])
        c.alignment = right_align
        c.border = thin_border

        # Total Duration
        c = ws.cell(row=row_idx, column=11, value=r["TotalDuration"])
        c.alignment = right_align
        c.border = thin_border
        
        c = ws.cell(row=row_idx, column=12, value=r["IPAddress"])
        c.border = thin_border
        
        c = ws.cell(row=row_idx, column=13, value=r["Note"])
        c.border = thin_border

        row_idx += 1

    # ----- Auto-size columns (avoid merged-cell issue) -----
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    wb.save(path)
    print(f"Wrote Excel report to {path}")
    return path

def get_device_groups():
    """
    Returns list of (group_id, group_name) from DeviceGroup.
    Uses exactly: SELECT nDeviceGroupID, sGroupName FROM DeviceGroup;
    """
    conn = pyodbc.connect(get_connection_string())
    cursor = conn.cursor()
    cursor.execute("SELECT nDeviceGroupID, sGroupName FROM DeviceGroup")
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return [(int(row[0]), row[1]) for row in rows]

def get_previous_month_range():
    now = datetime.now()
    first_this_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    if first_this_month.month == 1:
        first_prev_month = first_this_month.replace(year=first_this_month.year - 1, month=12)
    else:
        first_prev_month = first_this_month.replace(month=first_this_month.month - 1)

    # start = first day of previous month, end = first day of this month
    return first_prev_month, first_this_month


def run_scheduled_reports():
    """
    Execute all due scheduled reports based on the Excel configuration
    and persisted state. This is safe to call repeatedly (e.g. from a
    background scheduler); it will only generate reports that are due.
    """
    # 1) Load config (JSON schedule) and state (JSON)
    # jobs: list of schedule rows (each row is independent, even if group repeats)
    jobs = load_schedule_config()
    state = load_state()

    # Prune state entries for deleted jobs.
    # If a schedule row is removed, its job id won't exist anymore,
    # so we delete its last-run markers from state.json.
    valid_job_ids = set()
    for idx, job in enumerate(jobs):
        group_name = (job.get("group") or "").strip()
        jid = (job.get("id") or "").strip() or f"{group_name}::__idx_{idx}"
        if jid:
            valid_job_ids.add(jid)

    if isinstance(state, dict):
        keys_to_delete = []
        for k in state.keys():
            if k == "__MONTHLY_ALL__":
                continue
            if "::__" not in k:
                continue
            prefix = k.split("::__", 1)[0]
            if prefix not in valid_job_ids:
                keys_to_delete.append(k)

        for k in keys_to_delete:
            del state[k]
    # --- Monthly all-groups run (once per month) ---
    last_monthly_str = state.get("__MONTHLY_ALL__")
    now = _scheduler_now()
    do_monthly = False

    if last_monthly_str is None:
        # never ran before -> run now
        do_monthly = True
    else:
        last_monthly = datetime.fromisoformat(last_monthly_str)
        if last_monthly.year != now.year or last_monthly.month != now.month:
            # last run was in a previous month -> run now
            do_monthly = True

    if do_monthly:
        print("[MONTHLY] Generating all-group monthly reports...")
        start_m, end_m = get_previous_month_range()

        # generate reports for ALL groups for last month
        groups = get_device_groups()  # (id, name)
        for gid, gname in groups:
            print(f"[MONTHLY RUN] {gname}: {start_m} -> {end_m}")
            try:
                write_excel_for_group(
                    device_group_id=gid,
                    start_date=start_m,
                    end_date=end_m,
                    group_name=gname,
                )
            except Exception as e:
                print(f"[MONTHLY ERROR] {gname}: {e}")

        # send everything as one big mail
        send_all_reports_via_email(start_m, end_m)

        # update monthly marker
        state["__MONTHLY_ALL__"] = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()

    # 2) Get groups from DB and build name -> id map
    groups = get_device_groups()        # returns list of (id, name)
    name_to_id = {name: gid for gid, name in groups}

    # 3) Trigger-based loop: one run per job when trigger fires, same window for both reports
    for idx, job in enumerate(jobs):
        group_name_stripped = (job.get("group") or "").strip()
        job_id = (job.get("id") or "").strip() or f"{group_name_stripped}::__idx_{idx}"

        if group_name_stripped not in name_to_id:
            print(f"[SKIP] Config group '{group_name_stripped}' not found in DeviceGroup table")
            continue
        st = (job.get("type") or "weekly").strip().lower()
        if st == "monthly" and job.get("run_day_of_month") is None:
            print(f"[SKIP] Job '{group_name_stripped}' (monthly) missing run_day_of_month")
            continue
        if st == "weekly" and not job.get("run_day"):
            print(f"[SKIP] Job '{group_name_stripped}' (weekly) missing run_day")
            continue
        if not job.get("run_time"):
            print(f"[SKIP] Job '{group_name_stripped}' missing run_time")
            continue

        state_key = f"{job_id}::__last_run"
        last_run_iso = state.get(state_key)
        should_run, anchor_dt = _trigger_fires(job, now, last_run_iso)
        if not should_run or anchor_dt is None:
            continue

        try:
            start_date, end_date = _compute_window_from_trigger(anchor_dt, job)
        except Exception as e:
            print(f"[SKIP] Window computation for '{group_name_stripped}': {e}")
            continue
        if start_date >= end_date:
            print(f"[SKIP] Invalid window for '{group_name_stripped}': start >= end")
            continue

        device_group_id = name_to_id[group_name_stripped]
        run_reports = (job.get("report_type") or "both").strip().lower()

        # Report 1: Active Monitor Availability (if requested)
        if run_reports in ("availability", "both"):
            try:
                print(f"[RUN] Availability for {group_name_stripped}: {start_date} -> {end_date}")
                availability_report_path = write_excel_for_group(
                    device_group_id=device_group_id,
                    start_date=start_date,
                    end_date=end_date,
                    group_name=group_name_stripped,
                )
                send_single_report_email(
                    group_name=group_name_stripped,
                    report_path=availability_report_path,
                    start_date=start_date,
                    end_date=end_date,
                )
            except Exception as e:
                print(f"[ERROR] Availability report for '{group_name_stripped}': {e}")

        # Report 2: Device Uptime (if requested)
        if run_reports in ("uptime", "both"):
            try:
                print(f"[RUN] DeviceUpTime for {group_name_stripped}: {start_date} -> {end_date}")
                uptime_rows = run_sp_group_device_uptime(
                    device_group_id=device_group_id,
                    start_date=start_date,
                    end_date=end_date,
                )
                if uptime_rows:
                    uptime_report_path = write_uptime_excel(
                        group_name_stripped,
                        uptime_rows,
                        start_date,
                        end_date,
                    )
                    send_single_report_email(
                        group_name=group_name_stripped,
                        report_path=uptime_report_path,
                        start_date=start_date,
                        end_date=end_date,
                    )
                else:
                    print(f"[RUN] No uptime data for '{group_name_stripped}' in window")
            except Exception as e:
                print(f"[ERROR] DeviceUpTime report for '{group_name_stripped}': {e}")

        state[state_key] = anchor_dt.replace(second=0, microsecond=0).isoformat()

    save_state(state)
