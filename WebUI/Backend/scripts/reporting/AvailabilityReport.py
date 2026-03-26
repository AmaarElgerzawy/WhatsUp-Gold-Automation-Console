"""
Active Monitor Availability report: fetch data from WUG DB and write Excel.
No scheduling or email logic; use ReportScheduler for that.
"""
import os
import pyodbc
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from constants import get_connection_string

OUTPUT_FOLDER = r"C:\WUG_Exports"
WEB_USER_ID = 1
REPORT_PERCENT_DIVISOR = 10000000.0
ROUND_TO_PLACES_EXPORT = 7


def get_duration_from_seconds(total_seconds: int) -> str:
    """Approximate of GetDurationFromSeconds from ASP code."""
    if total_seconds is None or total_seconds < 0:
        total_seconds = 0
    total_seconds = int(total_seconds)
    minutes, _ = divmod(total_seconds, 60)
    hours, minutes = divmod(minutes, 60)
    days, hours = divmod(hours, 24)
    parts = []
    if days:
        parts.append(f"{days}d")
    if hours:
        parts.append(f"{hours}h")
    if minutes or not parts:
        parts.append(f"{minutes}m")
    return " ".join(parts)


def get_active_monitor_availability(
    device_group_id: int,
    start_date: datetime,
    end_date: datetime,
):
    """
    Runs the same SQL as SSMS (LoadDeviceGroup + aggregation) and returns rows for export.
    """
    start_str = start_date.strftime("%Y-%m-%dT%H:%M:%S")
    end_str = end_date.strftime("%Y-%m-%dT%H:%M:%S")

    sql_template = f"""
DECLARE @DeviceGroupID INT;
DECLARE @WebUserID INT;
DECLARE @StartDate DATETIME;
DECLARE @EndDate DATETIME;

SET @DeviceGroupID = {device_group_id};
SET @WebUserID = {WEB_USER_ID};
SET @StartDate = '{start_str}';
SET @EndDate = '{end_str}';

DECLARE @PivotTableName SYSNAME;
SET @PivotTableName = 'PivotDeviceToGroupTemp';

IF OBJECT_ID(@PivotTableName) IS NOT NULL
BEGIN
    DECLARE @dropSql NVARCHAR(4000);
    SET @dropSql = N'DROP TABLE ' + QUOTENAME(@PivotTableName) + N';';
    EXEC (@dropSql);
END;

EXEC LoadDeviceGroup @DeviceGroupID, @WebUserID, @PivotTableName;

DECLARE @sql NVARCHAR(MAX);

SET @sql = N'
SELECT
    Device.nDeviceID,
    Device.sDisplayName,
    NetworkInterface.nNetworkInterfaceID,
    NetworkInterface.sNetworkName,
    NetworkInterface.sNetworkAddress,
    Device.nWorstStateID,
    Device.nBestStateID,
    ActiveMonitorType.sMonitorTypeName,
    PivotActiveMonitorTypeToDevice.sArgument,
    PivotActiveMonitorTypeToDevice.nPivotActiveMonitorTypeToDeviceID,
    PivotActiveMonitorTypeToDevice.sComment,
    CAST(Device.sNote AS NVARCHAR(MAX)) AS sNote,

    SUM(
        CASE MonitorState.nInternalMonitorState WHEN 1 THEN
            DATEDIFF(
                SECOND,
                CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
                ISNULL(
                    CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                    CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
                )
            )
        ELSE 0 END
    ) AS nDownSeconds,

    SUM(
        CASE MonitorState.nInternalMonitorState WHEN 2 THEN
            DATEDIFF(
                SECOND,
                CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
                ISNULL(
                    CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                    CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
                )
            )
        ELSE 0 END
    ) AS nMaintenanceSeconds,

    SUM(
        CASE MonitorState.nInternalMonitorState WHEN 3 THEN
            DATEDIFF(
                SECOND,
                CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
                ISNULL(
                    CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                    CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
                )
            )
        ELSE 0 END
    ) AS nUpSeconds,

    SUM(
        CASE MonitorState.nInternalMonitorState WHEN -1 THEN
            DATEDIFF(
                SECOND,
                CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
                ISNULL(
                    CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                    CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
                )
            )
        ELSE 0 END
    ) AS nUnknownSeconds,

    SUM(
        DATEDIFF(
            SECOND,
            CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
            ISNULL(
                CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
            )
        )
    ) AS nTotalSeconds,

    (1.0 * SUM(
        CASE MonitorState.nInternalMonitorState WHEN 1 THEN
            DATEDIFF(
                SECOND,
                CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
                ISNULL(
                    CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                    CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
                )
            )
        ELSE 0 END
    ) / NULLIF(
        SUM(
            DATEDIFF(
                SECOND,
                CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
                ISNULL(
                    CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                    CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
                )
            )
        ), 0)
    ) * 1000000000.0 AS nDownPercent,

    (1.0 * SUM(
        CASE MonitorState.nInternalMonitorState WHEN 2 THEN
            DATEDIFF(
                SECOND,
                CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
                ISNULL(
                    CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                    CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
                )
            )
        ELSE 0 END
    ) / NULLIF(
        SUM(
            DATEDIFF(
                SECOND,
                CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
                ISNULL(
                    CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                    CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
                )
            )
        ), 0)
    ) * 1000000000.0 AS nMaintenancePercent,

    (1.0 * SUM(
        CASE MonitorState.nInternalMonitorState WHEN 3 THEN
            DATEDIFF(
                SECOND,
                CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
                ISNULL(
                    CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                    CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
                )
            )
        ELSE 0 END
    ) / NULLIF(
        SUM(
            DATEDIFF(
                SECOND,
                CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
                ISNULL(
                    CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                    CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
                )
            )
        ), 0)
    ) * 1000000000.0 AS nUpPercent,

    (1.0 * SUM(
        CASE MonitorState.nInternalMonitorState WHEN -1 THEN
            DATEDIFF(
                SECOND,
                CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
                ISNULL(
                    CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                    CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
                )
            )
        ELSE 0 END
    ) / NULLIF(
        SUM(
            DATEDIFF(
                SECOND,
                CASE WHEN dStartTime < @StartDate THEN @StartDate ELSE dStartTime END,
                ISNULL(
                    CASE WHEN dEndTime > @EndDate THEN @EndDate ELSE dEndTime END,
                    CASE WHEN @EndDate < GETDATE() THEN @EndDate ELSE GETDATE() END
                )
            )
        ), 0)
    ) * 1000000000.0 AS nUnknownPercent

FROM ActiveMonitorStateChangeLog
INNER JOIN MonitorState
    ON MonitorState.nMonitorStateID = ActiveMonitorStateChangeLog.nMonitorStateID
INNER JOIN PivotActiveMonitorTypeToDevice
    ON PivotActiveMonitorTypeToDevice.nPivotActiveMonitorTypeToDeviceID = ActiveMonitorStateChangeLog.nPivotActiveMonitorTypeToDeviceID
INNER JOIN Device
    ON Device.nDeviceID = PivotActiveMonitorTypeToDevice.nDeviceID
INNER JOIN NetworkInterface
    ON NetworkInterface.nNetworkInterfaceID = Device.nDefaultNetworkInterfaceID
INNER JOIN ActiveMonitorType
    ON ActiveMonitorType.nActiveMonitorTypeID = PivotActiveMonitorTypeToDevice.nActiveMonitorTypeID
INNER JOIN ' + QUOTENAME(@PivotTableName) + N'
    ON Device.nDeviceID = ' + QUOTENAME(@PivotTableName) + N'.nDeviceID
WHERE
    ISNULL(ActiveMonitorType.bRemoved,0) <> 1
    AND ISNULL(Device.bRemoved,0)      <> 1
    AND (
        (dStartTime >= @StartDate AND ISNULL(dEndTime, GETDATE()) <= @EndDate)
        OR (dStartTime <= @EndDate   AND ISNULL(dEndTime, GETDATE()) >= @StartDate)
        OR (dStartTime <= @EndDate   AND ISNULL(dEndTime, GETDATE()) >= @EndDate)
        OR (dStartTime <= @StartDate AND ISNULL(dEndTime, GETDATE()) >= @EndDate)
    )
GROUP BY
    Device.nDeviceID,
    Device.sDisplayName,
    NetworkInterface.nNetworkInterfaceID,
    NetworkInterface.sNetworkName,
    NetworkInterface.sNetworkAddress,
    Device.nWorstStateID,
    Device.nBestStateID,
    PivotActiveMonitorTypeToDevice.nPivotActiveMonitorTypeToDeviceID,
    PivotActiveMonitorTypeToDevice.sArgument,
    PivotActiveMonitorTypeToDevice.sComment,
    ActiveMonitorType.sMonitorTypeName,
    CAST(Device.sNote AS NVARCHAR(MAX))
ORDER BY
    Device.sDisplayName ASC,
    ActiveMonitorType.sMonitorTypeName ASC;
';

EXEC sp_executesql
    @sql,
    N'@StartDate datetime, @EndDate datetime',
    @StartDate = @StartDate,
    @EndDate   = @EndDate;

IF OBJECT_ID(@PivotTableName) IS NOT NULL
BEGIN
    DECLARE @dropSql2 NVARCHAR(4000);
    SET @dropSql2 = N'DROP TABLE ' + QUOTENAME(@PivotTableName) + N';';
    EXEC (@dropSql2);
END;
"""

    conn = pyodbc.connect(get_connection_string())
    cursor = conn.cursor()
    cursor.execute(sql_template)
    rows = cursor.fetchall()
    cols = [c[0] for c in cursor.description]
    results = []

    for row in rows:
        rec = dict(zip(cols, row))

        def scaled(p):
            if p is None:
                return 0.0
            return float(p) / REPORT_PERCENT_DIVISOR

        up_pct = scaled(rec["nUpPercent"])
        down_pct = scaled(rec["nDownPercent"])
        maint_pct = scaled(rec["nMaintenancePercent"])
        unknown_pct = scaled(rec["nUnknownPercent"])

        up_dur = get_duration_from_seconds(rec["nUpSeconds"])
        down_dur = get_duration_from_seconds(rec["nDownSeconds"])
        maint_dur = get_duration_from_seconds(rec["nMaintenanceSeconds"])
        unknown_dur = get_duration_from_seconds(rec["nUnknownSeconds"])
        total_dur = get_duration_from_seconds(rec["nTotalSeconds"])

        sMonitorTypeName = rec["sMonitorTypeName"]
        sArgument = rec["sArgument"] or ""
        sComment = rec["sComment"] or ""
        sIPAddress = rec["sNetworkAddress"]
        sNote = rec["sNote"]

        monitor_name = sMonitorTypeName
        if sArgument:
            monitor_name += f" ({sArgument})"
        if sComment:
            monitor_name += f" - {sComment}"

        results.append({
            "Device": rec["sDisplayName"],
            "IPAddress": sIPAddress,
            "Note": sNote,
            "Monitor": monitor_name,
            "UpPercent": round(up_pct, ROUND_TO_PLACES_EXPORT),
            "UpDuration": up_dur,
            "MaintenancePercent": round(maint_pct, ROUND_TO_PLACES_EXPORT),
            "MaintenanceDuration": maint_dur,
            "UnknownPercent": round(unknown_pct, ROUND_TO_PLACES_EXPORT),
            "UnknownDuration": unknown_dur,
            "DownPercent": round(down_pct, ROUND_TO_PLACES_EXPORT),
            "DownDuration": down_dur,
            "TotalDuration": total_dur,
        })

    cursor.close()
    conn.close()
    return results


def write_excel_for_group(
    device_group_id: int,
    start_date: datetime,
    end_date: datetime,
    group_name: str = None,
):
    rows = get_active_monitor_availability(device_group_id, start_date, end_date)
    if not rows:
        print(f"No data for group {device_group_id}")
        return None

    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    if not group_name:
        group_name = f"group_{device_group_id}"

    safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in group_name)
    filename = f"ActiveMonitorAvailability_{safe_name}.xlsx"
    path = os.path.join(OUTPUT_FOLDER, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Active Monitor Availability"

    headers = [
        "Device",
        "Monitor",
        "Up",
        "Up Duration",
        "Maintenance",
        "Maintenance Duration",
        "Unknown",
        "Unknown Duration",
        "Down",
        "Down Duration",
        "Total Duration",
        "Network Address (IP)",
        "Note",
    ]
    last_col_letter = get_column_letter(len(headers))

    title_font = Font(size=14, bold=True)
    subtitle_font = Font(size=11)
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center")
    right_align = Alignment(horizontal="right")
    header_fill = PatternFill("solid", fgColor="C0C0C0")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws["A1"].value = "Active Monitor Availability"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align
    ws.merge_cells(f"A1:{last_col_letter}1")

    start_long = start_date.strftime("%A, %B %d, %Y %I:%M:%S %p")
    end_long = end_date.strftime("%A, %B %d, %Y %I:%M:%S %p")
    ws["A2"].value = f"{group_name} - {start_long} - {end_long}"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = center_align
    ws.merge_cells(f"A2:{last_col_letter}2")

    header_row_idx = 3
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    row_idx = header_row_idx + 1
    for r in rows:
        up_fraction = (r["UpPercent"] or 0.0) / 100.0
        maint_fraction = (r["MaintenancePercent"] or 0.0) / 100.0
        unknown_fraction = (r["UnknownPercent"] or 0.0) / 100.0
        down_fraction = (r["DownPercent"] or 0.0) / 100.0

        ws.cell(row=row_idx, column=1, value=r["Device"]).border = thin_border
        ws.cell(row=row_idx, column=2, value=r["Monitor"]).border = thin_border

        up_cell = ws.cell(row=row_idx, column=3, value=up_fraction)
        up_cell.number_format = "0.0000000%"
        up_cell.alignment = right_align
        up_cell.border = thin_border

        ws.cell(row=row_idx, column=4, value=r["UpDuration"]).alignment = right_align
        ws.cell(row=row_idx, column=4).border = thin_border

        maint_cell = ws.cell(row=row_idx, column=5, value=maint_fraction)
        maint_cell.number_format = "0.0000000%"
        maint_cell.alignment = right_align
        maint_cell.border = thin_border

        ws.cell(row=row_idx, column=6, value=r["MaintenanceDuration"]).alignment = right_align
        ws.cell(row=row_idx, column=6).border = thin_border

        unknown_cell = ws.cell(row=row_idx, column=7, value=unknown_fraction)
        unknown_cell.number_format = "0.0000000%"
        unknown_cell.alignment = right_align
        unknown_cell.border = thin_border

        ws.cell(row=row_idx, column=8, value=r["UnknownDuration"]).alignment = right_align
        ws.cell(row=row_idx, column=8).border = thin_border

        down_cell = ws.cell(row=row_idx, column=9, value=down_fraction)
        down_cell.number_format = "0.0000000%"
        down_cell.alignment = right_align
        down_cell.border = thin_border

        ws.cell(row=row_idx, column=10, value=r["DownDuration"]).alignment = right_align
        ws.cell(row=row_idx, column=10).border = thin_border

        ws.cell(row=row_idx, column=11, value=r["TotalDuration"]).alignment = right_align
        ws.cell(row=row_idx, column=11).border = thin_border

        ws.cell(row=row_idx, column=12, value=r["IPAddress"]).border = thin_border
        ws.cell(row=row_idx, column=13, value=r["Note"]).border = thin_border

        row_idx += 1

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    wb.save(path)
    print(f"Wrote Excel report to {path}")
    return path


def get_device_groups():
    """Returns list of (group_id, group_name) from DeviceGroup."""
    conn = pyodbc.connect(get_connection_string())
    cursor = conn.cursor()
    cursor.execute("SELECT nDeviceGroupID, sGroupName FROM DeviceGroup")
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return [(int(row[0]), row[1]) for row in rows]
