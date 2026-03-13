import os
from pandas import pivot_table
import pyodbc
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from constants import get_connection_string

# ================= CONFIG =================

OUTPUT_FOLDER = r"C:\WUG_Exports"
WEB_USER_ID = 1

REPORT_PERCENT_DIVISOR = 10_000_000.0
ROUND_TO_PLACES_EXPORT = 7

# =========================================
def get_duration_from_seconds(total_seconds: int) -> str:
    if not total_seconds or total_seconds < 0:
        total_seconds = 0

    minutes, _ = divmod(int(total_seconds), 60)
    hours, minutes = divmod(minutes, 60)
    days, hours = divmod(hours, 24)

    parts = []
    if days:
        parts.append(f"{days}d")
    if hours:
        parts.append(f"{hours}h")
    if minutes or not parts:
        parts.append(f"{minutes}m")

    return " ".join(parts)

def get_device_extra_data(device_ids):

    conn = pyodbc.connect(get_connection_string())
    cur = conn.cursor()

    ids = ",".join(str(i) for i in device_ids)

    sql = f"""
    SELECT
        d.nDeviceID,
        d.sNote,
        ni.sNetworkAddress AS InterfaceAddress
    FROM Device d
    LEFT JOIN NetworkInterface ni
        ON ni.nDeviceID = d.nDeviceID
    WHERE d.nDeviceID IN ({ids})
    """

    cur.execute(sql)

    data = {}
    for row in cur.fetchall():
        data[row.nDeviceID] = {
            "Note": row.sNote,
            "InterfaceAddress": row.InterfaceAddress
        }

    cur.close()
    conn.close()

    return data

def run_sp_group_device_uptime(
    device_group_id: int,
    start_date: datetime,
    end_date: datetime
):
    """
    Calls spGroupDeviceUptime exactly like the ASP code and
    returns the final result set.
    """

    start_str = start_date.strftime("%Y-%m-%dT%H:%M:%S")
    end_str = end_date.strftime("%Y-%m-%dT%H:%M:%S")

    pivot_table = "PivotDeviceToGroup_Python"

    sql = f"""
DECLARE @PivotTable SYSNAME = '{pivot_table}';

IF OBJECT_ID(@PivotTable,'U') IS NOT NULL
BEGIN
    EXEC('DROP TABLE ' + @PivotTable);
END;

EXEC dbo.spGroupDeviceUptime
    @nDeviceGroupID      = {device_group_id},
    @dSqlStartDate       = '{start_str}',
    @dSqlEndDate         = '{end_str}',
    @nWebUserId          = {WEB_USER_ID},
    @sPivotDeviceToGroupTable = '{pivot_table}',
    @sSortBy             = 'sDisplayName',
    @sSortDirection      = 'ASC',
    @nLowCount           = 1,
    @nHighCount          = 32767,
    @nRowCount           = '32767',
    @bAllPages           = 1;

IF OBJECT_ID(@PivotTable,'U') IS NOT NULL
BEGIN
    EXEC('DROP TABLE ' + @PivotTable);
END;
"""

    conn = pyodbc.connect(get_connection_string())
    cur = conn.cursor()
    cur.execute(sql)

    cols = [c[0] for c in cur.description]
    rows = cur.fetchall()

    cur.close()
    conn.close()

    results = []

    for row in rows:
        rec = dict(zip(cols, row))

        def pct(v):
            return round((v or 0) / REPORT_PERCENT_DIVISOR, ROUND_TO_PLACES_EXPORT)

        results.append({
            "DeviceID": rec["nDeviceID"],
            "Device": rec["sDisplayName"],
            "Address": rec["sNetworkAddress"],
            "Up": pct(rec["nUpPercent"]),
            "UpDuration": get_duration_from_seconds(rec["nUpSeconds"]),
            "Maintenance": pct(rec["nMaintenancePercent"]),
            "MaintenanceDuration": get_duration_from_seconds(rec["nMaintenanceSeconds"]),
            "Unknown": pct(rec["nUnknownPercent"]),
            "UnknownDuration": get_duration_from_seconds(rec["nUnknownSeconds"]),
            "Down": pct(rec["nDownPercent"]),
            "DownDuration": get_duration_from_seconds(rec["nDownSeconds"]),
            "TotalDuration": get_duration_from_seconds(rec["nTotalSeconds"]),
        })

        device_ids = [r["DeviceID"] for r in results]
        if not device_ids:
            return results

        extra_data = get_device_extra_data(device_ids)

        for r in results:
            extra = extra_data.get(r["DeviceID"], {})
            r["Note"] = extra.get("Note")
    return results

def write_excel(group_name, rows, start_date, end_date):
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in group_name)
    path = os.path.join(
        OUTPUT_FOLDER,
        f"DeviceUpTime_{safe_name}.xlsx"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Active Monitor Availability"

    headers = [
        "Device",
        "Address",
        "Note",
        "Up", "Up Duration",
        "Maintenance", "Maintenance Duration",
        "Unknown", "Unknown Duration",
        "Down", "Down Duration",
        "Total Duration"
    ]

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")
    fill = PatternFill("solid", fgColor="C0C0C0")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"] = "Active Monitor Availability"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = center

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A2"] = f"{group_name} - {start_date} → {end_date}"
    ws["A2"].alignment = center

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = bold
        c.fill = fill
        c.alignment = center
        c.border = border

    r_idx = 4
    for r in rows:
        ws.cell(r_idx, 1, r["Device"]).border = border
        ws.cell(r_idx, 2, r["Address"]).border = border
        ws.cell(r_idx, 3, r["InterfaceAddress"]).border = border
        ws.cell(r_idx, 4, r["Note"]).border = border

        for col, key in [(5,"Up"),(7,"Maintenance"),(9,"Unknown"),(11,"Down")]:
            c = ws.cell(r_idx, col, r[key] / 100)
            c.number_format = "0.0000000%"
            c.alignment = right
            c.border = border

        ws.cell(r_idx, 6, r["UpDuration"])
        ws.cell(r_idx, 8, r["MaintenanceDuration"])
        ws.cell(r_idx, 10, r["UnknownDuration"])
        ws.cell(r_idx, 12, r["DownDuration"])
        ws.cell(r_idx, 13, r["TotalDuration"])

        r_idx += 1

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22

    wb.save(path)
    print(f"Excel written: {path}")
    return path
